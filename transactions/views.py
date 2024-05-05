from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph,Spacer
from io import BytesIO
from reportlab.lib.styles import getSampleStyleSheet

from .models import Transactions
from .serializers import TransactionSerializer

# Create your views here.

class TransactionList(APIView):
    def get(self,request):
        try:
            transactions = Transactions.objects.all()
            serializer = TransactionSerializer(transactions, many=True)
            return Response({"message":"checking",'data':serializer.data})
        except Exception as e:
            return Response({"message": f'Something went wrong with {str(e)}',
                             'status': status.HTTP_500_INTERNAL_SERVER_ERROR})
        

class ExcelGenerator(APIView):
    def get(self,request):
        try:
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("start_date")
            company = request.GET.get("company")
            project = request.GET.get("project",None)
            ledger = request.GET.get("ledger")
            if project:
                transactions = Transactions.objects.filter(
                    Q(payer=ledger) | Q(receiver=ledger),
                    company=company,
                    created_at__date__gte=start_date,
                    created_at__date__lte=end_date,
                    project=project
                    )
            else:
                transactions = Transactions.objects.filter(
                    Q(payer=ledger) | Q(receiver=ledger),
                    company=company,
                    created_at__date__gte=start_date,
                    created_at__date__lte=end_date,
                    )
            # Create Excel workbook and sheet
            wb = Workbook()
            ws = wb.active

            # Define headers
            headers = ['Serial No.', 'Head', 'Debit','Credit','Date']

            for col, header in enumerate(headers, start=1):
                ws.cell(row=2, column=col, value=header)
                ws.cell(row=2, column=col).font = Font(size=12)  # Set font size for column names
            
            
            # Write transactions data to Excel
            row_num = 3
            for idx, transaction in enumerate(transactions, start=1):
                ws.cell(row=row_num, column=1, value=idx).font = Font(size=9) # Serial No.
                if ledger == transaction.payer:
                    ws.cell(row=row_num, column=2, value=transaction.receiver).font = Font(size=9)
                elif ledger == transaction.receiver:
                    ws.cell(row=row_num, column=2, value=transaction.payer).font = Font(size=9)
                
                if ledger == transaction.payer:
                   ws.cell(row=row_num, column=3, value=transaction.amount).font = Font(size=9)
                elif ledger == transaction.receiver:
                   ws.cell(row=row_num, column=4, value=transaction.amount).font = Font(size=9)

                ws.cell(row=row_num, column=5, value=transaction.transaction_date).font = Font(size=9)
                row_num += 1
            

            heading = ledger
            sub_heading = project if project else 'N/A'
            opening_balance = 100
            current_balance = 56

            ws.merge_cells('A1:E1')  # Merge cells for the main heading
            main_heading_cell = ws['A1']
            if sub_heading == 'N/A':
               main_heading_cell.value = f"{heading} Petty Cash\nOpening Balance: {opening_balance}\nCurrent Balance: {current_balance}"
            else:
               main_heading_cell.value = f"{heading} Petty Cash\nProject: {sub_heading}\nOpening Balance: {opening_balance}\nCurrent Balance: {current_balance}"

            main_heading_cell.alignment = Alignment(horizontal='center')
            main_heading_cell.font = Font(size=20,bold=True)  # Set font size for main heading

            # Save the Excel file
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{ledger}-{start_date}-{end_date}.xlsx"'
            wb.save(response)

            return response
        except Exception as e:
            return Response({"message": f'Something went wrong with {str(e)}',
                                'status': status.HTTP_500_INTERNAL_SERVER_ERROR})




class PdfGenerator(APIView):
    def get(self, request):
        try:
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("end_date")  # Corrected typo
            company = request.GET.get("company")
            project = request.GET.get("project", None)
            ledger = request.GET.get("ledger")

            if project:
                transactions = Transactions.objects.filter(
                    Q(payer=ledger) | Q(receiver=ledger),
                    company=company,
                    created_at__date__gte=start_date,
                    created_at__date__lte=end_date,
                    project=project
                )
            else:
                transactions = Transactions.objects.filter(
                    Q(payer=ledger) | Q(receiver=ledger),
                    company=company,
                    created_at__date__gte=start_date,
                    created_at__date__lte=end_date,
                )

            # Create a PDF document
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)

            # Get the height of the document
   

            # Define a custom layout for the PDF document
            elements = []

                # Add heading and subheading
            heading_style = getSampleStyleSheet()["Heading1"]
            heading = Paragraph(f'{ledger} Petty Cash', heading_style)
            elements.append(heading)
            sub_heading_style = getSampleStyleSheet()["Heading2"]
            if project:
                sub_heading = Paragraph(f'Project: {project}', sub_heading_style)
                elements.append(sub_heading)
            else:
                sub_heading = ""
            balance_style = getSampleStyleSheet()["Heading3"]
            opening_balance = Paragraph(f'Opening Balance: {100}', balance_style)
            current_balance = Paragraph(f'Current Balance: {56}', balance_style)
            elements.extend([opening_balance,current_balance])

            # Calculate remaining space available on the page
            remaining_space = doc.height - sum(e.wrap(doc.width, doc.height)[1] for e in elements)

            # Add spacer with adjusted size
            spacer_height = min(remaining_space, 100)  # Limit the spacer height to 100
            elements.append(Spacer(1, spacer_height))

            # Define table headers
            headers = ['Serial No.', 'Head', 'Debit', 'Credit', 'Date']
            data = [headers]

            # Populate transaction data
            for idx, transaction in enumerate(transactions, start=1):
                row = [
                    idx,
                    transaction.receiver if ledger == transaction.payer else transaction.payer,
                    transaction.amount if ledger == transaction.payer else '',
                    transaction.amount if ledger == transaction.receiver else '',
                    str(transaction.transaction_date)
                ]
                data.append(row)

            # Create a table and style it
            table = Table(data)
            style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)])
            table.setStyle(style)

            elements.append(table)

            # Build PDF document
            doc.build(elements)

            # Get PDF data from buffer
            pdf_data = buffer.getvalue()
            buffer.close()

            # Return PDF as response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{ledger}-{start_date}-{end_date}.pdf"'
            response.write(pdf_data)

            return response

        except Exception as e:
            return Response({"message": f'Something went wrong with {str(e)}',
                             'status': status.HTTP_500_INTERNAL_SERVER_ERROR})
        